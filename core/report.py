"""
报告生成器

生成Markdown格式的分析报告。
"""

import os
import pandas as pd
from datetime import datetime
from typing import Dict, Any, List
import logging

logger = logging.getLogger(__name__)


class ReportGenerator:
    """报告生成器"""
    
    def __init__(self, base_dir: str):
        """
        初始化报告生成器
        
        Args:
            base_dir: 基础目录路径
        """
        self.base_dir = base_dir
        self.reports_dir = os.path.join(base_dir, "outputs", "reports")
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_daily_report(self, analysis_result: Dict[str, Any], 
                            recent_scores: pd.DataFrame = None) -> str:
        """
        生成每日分析报告
        
        Args:
            analysis_result: 分析结果
            recent_scores: 最近评分数据
            
        Returns:
            报告文件路径
        """
        date = analysis_result["date"]
        date_str = date.strftime("%Y-%m-%d")
        
        # 构建报告内容
        content = self._build_report_content(analysis_result, recent_scores)
        
        # 保存报告
        report_file = f"macro_analysis_{date_str}.md"
        report_path = os.path.join(self.reports_dir, report_file)
        
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"生成报告: {report_path}")
        return report_path
    
    def _build_report_content(self, analysis_result: Dict[str, Any], 
                            recent_scores: pd.DataFrame = None) -> str:
        """
        构建报告内容
        
        Args:
            analysis_result: 分析结果
            recent_scores: 最近评分数据
            
        Returns:
            报告内容
        """
        date = analysis_result["date"]
        date_str = date.strftime("%Y年%m月%d日")
        
        factor_scores = analysis_result["factor_scores"]
        factor_values = analysis_result["factor_values"]
        total_score = analysis_result["total_score"]
        risk_level = analysis_result["risk_level"]
        
        # 报告标题
        content = f"# 宏观金融危机风险打分系统分析报告\n\n"
        content += f"**分析日期**: {date_str}\n\n"
        
        # 总体风险评分
        content += f"## 总体风险评分\n\n"
        content += f"**综合加权总分**: {total_score:.2f}\n\n"
        content += f"**风险等级**: {risk_level}\n\n"
        
        # 因子详细分析
        content += "## 因子详细分析\n\n"
        content += "| 因子名称 | 原始值 | 风险评分 | 权重 | 加权评分 |\n"
        content += "|---------|--------|----------|------|----------|\n"
        
        # 按权重排序显示因子
        sorted_factors = sorted(factor_scores.items(), 
                              key=lambda x: analysis_result.get("weights", {}).get(x[0], 0), 
                              reverse=True)
        
        for factor_id, score in sorted_factors:
            value = factor_values.get(factor_id, "N/A")
            weight = analysis_result.get("weights", {}).get(factor_id, 0)
            weighted_score = score * weight
            
            if isinstance(value, float):
                value_str = f"{value:.4f}"
            else:
                value_str = str(value)
            
            content += f"| {factor_id} | {value_str} | {score:.2f} | {weight:.2f} | {weighted_score:.2f} |\n"
        
        content += "\n"
        
        # 最近趋势
        if recent_scores is not None and not recent_scores.empty:
            content += "## 最近趋势\n\n"
            content += "| 日期 | 总分 |\n"
            content += "|------|------|\n"
            
            for _, row in recent_scores.iterrows():
                row_date = row['date'].strftime("%Y-%m-%d")
                content += f"| {row_date} | {row['total_score']:.2f} |\n"
            
            content += "\n"
        
        # 风险等级说明
        content += "## 风险等级说明\n\n"
        content += "- **低风险** (< 30): 市场风险较低，各项指标正常\n"
        content += "- **中等风险** (30-50): 市场存在一定风险，需要关注\n"
        content += "- **偏高风险** (50-70): 市场风险较高，建议谨慎\n"
        content += "- **极高风险** (≥ 70): 市场风险极高，建议高度警惕\n\n"
        
        # 报告生成时间
        content += f"---\n\n"
        content += f"*报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n"
        
        return content
    
    def generate_excel_summary(self, analysis_result: Dict[str, Any], 
                             excel_path: str) -> None:
        """
        生成Excel汇总报告
        
        Args:
            analysis_result: 分析结果
            excel_path: Excel文件路径
        """
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 检查文件是否存在
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
                # 删除默认工作表
                wb.remove(wb.active)
            
            # 创建工作表
            ws_name = analysis_result["date"].strftime("%Y-%m-%d")
            if ws_name in wb.sheetnames:
                ws = wb[ws_name]
            else:
                ws = wb.create_sheet(ws_name)
            
            # 设置标题行
            headers = [
                "日期", "VIX", "VIX分数", "SOFR原值", "3M国债", "TED利差", "TED分数",
                "高收益利差", "HY分数", "2Y-10Y利差", "Yield分数", "FCI", "FCI分数",
                "SP500波动率", "SP500分数", "美元指数波动", "DXY分数",
                "消费者信心", "消费者信心分数", "房价变化", "房价分数",
                "新兴市场利差", "EM风险分数", "加权总分"
            ]
            
            # 写入标题
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # 准备数据行
            date = analysis_result["date"]
            factor_scores = analysis_result["factor_scores"]
            factor_values = analysis_result["factor_values"]
            total_score = analysis_result["total_score"]
            
            # 映射因子ID到列位置
            factor_mapping = {
                "VIX": (1, 2),  # VIX, VIX分数
                "TED": (3, 6),  # SOFR原值, 3M国债, TED利差, TED分数
                "HY_Spread": (7, 8),  # 高收益利差, HY分数
                "Yield_Spread": (9, 10),  # 2Y-10Y利差, Yield分数
                "FCI": (11, 12),  # FCI, FCI分数
                "SP500_Vol": (13, 14),  # SP500波动率, SP500分数
                "DXY_Vol": (15, 16),  # 美元指数波动, DXY分数
                "Consumer_Confidence": (17, 18),  # 消费者信心, 消费者信心分数
                "Housing_Stress": (19, 20),  # 房价变化, 房价分数
                "EM_Risk": (21, 22),  # 新兴市场利差, EM风险分数
            }
            
            # 写入数据
            row = 2
            ws.cell(row=row, column=1, value=date.strftime("%Y-%m-%d"))
            
            for factor_id, (value_col, score_col) in factor_mapping.items():
                value = factor_values.get(factor_id, "")
                score = factor_scores.get(factor_id, 0)
                
                if isinstance(value, float):
                    ws.cell(row=row, column=value_col, value=round(value, 4))
                else:
                    ws.cell(row=row, column=value_col, value=value)
                
                ws.cell(row=row, column=score_col, value=round(score, 2))
            
            # 写入总分
            ws.cell(row=row, column=23, value=round(total_score, 2))
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 12
            
            # 保存文件
            wb.save(excel_path)
            logger.info(f"生成Excel报告: {excel_path}")
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
